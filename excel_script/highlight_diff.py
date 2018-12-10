import openpyxl as ox
import sys
import os


def read_data_from_excel(filename):
    """
    Функция читает одностраничные и многостраничные файлы Excel
    :param filename: путь к файлу, который хотим прочитать
    :return: возвращает head - "шапку" (названия столбцов, list); data - данные таблицы (list)
    """
    document = ox.load_workbook(filename=filename)
    sheets = document.sheetnames
    head = [cell.value for cell in list(document[sheets[0]].rows)[0]]
    data = []
    for sheet in sheets:
        data += [[cell.value for cell in row] for row in document[sheet].rows][1:]

    return head, data


def data_difference(head, source_file, updated_file):
    """
    Функция выявляет разницу между двумя файлами Excel: неизмененные строки остаются без цвета,
    добавленные во втором файле строки окрашиваются в зеленый цвет, удаленные во втором файле
    строки окрашиваются в красный цвет, строки второго файла, где были изменены данные (кроме ID)
    окрашиваются в желтый цвет
    Данные двух файлов сортируются по столбцу ID, далее применяем сортировку слиянием - сравниваем
    поэлементно два массива: если ID элемента первого массива меньше ID элемента второго массива, то
    строка первого массива была удалена (красим в красный), если ID равны - сравниваем содержание
    (одинаковое - не красим, разное - красим в желтый), если ID элемента первого массива больше ID
    элемента второго массива, то строка второго массива была добавлена (красим в зеленый). Если в
    одном из массивов мы не доходим до конца, то оставшиеся элементы добавляются в новый файл и
    красятся в нужный цвет (зеленый - если добавляются из второго массива, красный - из первого).
    :param head: шапка изначального документа
    :param source_file: данные изначального документа
    :param updated_file: данные измененного документа
    :return: data - данные для нового объединенного файла (list), colors - ключи для раскрашивания
    строк таблицы (list)
    """
    for index, value in enumerate(head):
        if value == "ID":
            id_index = index
    data_source_sorted = sorted(source_file, key=lambda column: column[id_index])
    data_updated_sorted = sorted(updated_file, key=lambda column: column[id_index])
    colors = []
    data = []

    i = 0
    j = 0
    while i < len(data_source_sorted) and j < len(data_updated_sorted):
        if data_source_sorted[i][id_index] < data_updated_sorted[j][id_index]:
            data.append(data_source_sorted[i])
            colors.append("red")
            i += 1
        elif data_source_sorted[i][id_index] == data_updated_sorted[j][id_index]:
            if data_source_sorted[i] == data_updated_sorted[j]:
                data.append(data_source_sorted[i])
                colors.append("")
            else:
                data.append(data_updated_sorted[j])
                colors.append("yellow")
            i += 1
            j += 1
        elif data_source_sorted[i][id_index] > data_updated_sorted[j][id_index]:
            data.append(data_updated_sorted[j])
            colors.append("green")
            j += 1

    while i < len(data_source_sorted):
        data.append(data_source_sorted[i])
        colors.append("red")
        i += 1
    while j < len(data_updated_sorted):
        data.append(data_updated_sorted[j])
        colors.append("green")
        j += 1

    return data, colors


def write_data_to_excel(filename, head, data, colors):
    """
    Функция записывает шапку и данные в файл Excel, применяя к ним окраску из массива colors
    :param filename: путь к новому файлу
    :param head: шапка таблицы (list)
    :param data: данные таблицы (list)
    :param colors: массив с цветами строк (list)
    """
    red = ox.styles.colors.Color(rgb='DA9694')
    red_fill = ox.styles.fills.PatternFill(patternType='solid', fgColor=red)
    yellow = ox.styles.colors.Color(rgb='FFFF00')
    yellow_fill = ox.styles.fills.PatternFill(patternType='solid', fgColor=yellow)
    green = ox.styles.colors.Color(rgb='C4D79B')
    green_fill = ox.styles.fills.PatternFill(patternType='solid', fgColor=green)

    wb = ox.Workbook()
    ws = wb.active
    data = [head] + data
    colors = [""] + colors
    for row in range(len(data)):
        for col in range(len(data[row])):
            _ = ws.cell(column=col + 1, row=row + 1, value=data[row][col])
            if colors[row] == "red":
                _.fill = red_fill
            if colors[row] == "yellow":
                _.fill = yellow_fill
            if colors[row] == "green":
                _.fill = green_fill
    wb.save(filename)


if __name__ == "__main__":
    source = os.path.abspath(sys.argv[1])
    updated = os.path.abspath(sys.argv[2])
    new = os.path.abspath(sys.argv[3])

    head, data_source = read_data_from_excel(source)
    useless_head, data_updated = read_data_from_excel(updated)

    data, colors = data_difference(head, data_source, data_updated)

    write_data_to_excel(new, head, data, colors)
